"""
Client Microsoft Graph API.

Responsabilità:
- Autenticazione via MSAL (token cache + browser interattivo)
- Download email con CORPO COMPLETO (non solo bodyPreview)
- Stripping HTML → testo pulito per l'embedding
- Download e estrazione testo dagli allegati (PDF, Word, Excel)
- Conversione dict grezzo → dataclass Email

Il resto del sistema vede solo list[Email].
"""

import base64
import io
import logging
import re

import requests
from bs4 import BeautifulSoup
from msal import PublicClientApplication

from src.config.settings import settings
from src.models.email import Email

logger = logging.getLogger(__name__)

_GRAPH_SCOPES = ["https://graph.microsoft.com/Mail.Read"]

# Chiede body completo (HTML) invece del solo bodyPreview
_MESSAGES_URL = (
    "https://graph.microsoft.com/v1.0/me/messages"
    "?$select=id,subject,from,toRecipients,receivedDateTime,"
    "body,bodyPreview,conversationId,hasAttachments"
    "&$top=50"
    "&$orderby=receivedDateTime desc"
)

# MIME type supportati per l'estrazione testo allegati
_SUPPORTED_ATTACHMENT_TYPES = {
    "application/pdf",
    "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "application/msword",
    "text/plain",
}


class GraphClient:
    def __init__(self) -> None:
        self._app = PublicClientApplication(
            settings.azure_client_id,
            authority=f"https://login.microsoftonline.com/{settings.azure_tenant_id}",
        )
        self._token: str | None = None

    # ── Autenticazione ───────────────────────────────────────────────────────

    def authenticate(self) -> str:
        if self._token:
            return self._token

        accounts = self._app.get_accounts()
        if accounts:
            result = self._app.acquire_token_silent(_GRAPH_SCOPES, account=accounts[0])
            if result and "access_token" in result:
                logger.info("Token recuperato dalla cache.")
                self._token = result["access_token"]
                return self._token

        logger.info("Apertura browser per autenticazione Microsoft...")
        result = self._app.acquire_token_interactive(scopes=_GRAPH_SCOPES)
        if "access_token" not in result:
            raise RuntimeError(
                f"Autenticazione fallita: {result.get('error_description', result)}"
            )
        logger.info("Autenticazione completata.")
        self._token = result["access_token"]
        return self._token

    # ── Fetch email ──────────────────────────────────────────────────────────

    def fetch_emails(self, max_emails: int | None = None) -> list[Email]:
        limit = max_emails or settings.max_emails
        token = self.authenticate()
        headers = {
            "Authorization": f"Bearer {token}",
            "Content-Type": "application/json",
        }

        raw_emails: list[dict] = []
        url: str | None = _MESSAGES_URL

        while url and len(raw_emails) < limit:
            logger.info("Download... %d/%d email", len(raw_emails), limit)
            resp = requests.get(url, headers=headers)
            if not resp.ok:
                logger.error("Errore Graph API (%d): %s", resp.status_code, resp.text)
            resp.raise_for_status()

            data = resp.json()
            raw_emails.extend(data.get("value", []))
            url = data.get("@odata.nextLink")

        raw_emails = raw_emails[:limit]
        logger.info("Scaricate %d email.", len(raw_emails))

        emails = []
        for raw in raw_emails:
            email = self._parse(raw)

            # Scarica allegati se l'email li ha e la feature è attiva
            if settings.fetch_attachments and raw.get("hasAttachments"):
                att_text = self._fetch_attachments(raw["id"], headers)
                if att_text:
                    email.attachments_text = att_text

            emails.append(email)

        return emails

    # ── Allegati ─────────────────────────────────────────────────────────────

    def _fetch_attachments(self, message_id: str, headers: dict) -> str:
        """
        Scarica gli allegati di un messaggio e ne estrae il testo.
        Supporta PDF, Word (.docx), Excel (.xlsx) e testo semplice.
        Ritorna il testo combinato di tutti gli allegati o stringa vuota.
        """
        url = (
            f"https://graph.microsoft.com/v1.0/me/messages/{message_id}/attachments"
            "?$select=name,contentType,contentBytes"
        )
        try:
            resp = requests.get(url, headers=headers)
            if not resp.ok:
                return ""
            attachments = resp.json().get("value", [])
        except Exception:
            return ""

        extracted: list[str] = []
        for att in attachments:
            content_type = att.get("contentType", "")
            content_bytes = att.get("contentBytes")
            name = att.get("name", "allegato")

            if not content_bytes or content_type not in _SUPPORTED_ATTACHMENT_TYPES:
                continue

            try:
                raw_bytes = base64.b64decode(content_bytes)
                text = self._extract_text(raw_bytes, content_type, name)
                if text and text.strip():
                    extracted.append(f"[{name}]\n{text.strip()}")
            except Exception as exc:
                logger.warning("Errore estrazione allegato '%s': %s", name, exc)

        return "\n\n".join(extracted)

    @staticmethod
    def _extract_text(data: bytes, content_type: str, name: str) -> str:
        """Estrae testo da un allegato in base al suo MIME type."""
        if content_type == "text/plain":
            return data.decode("utf-8", errors="replace")

        if content_type == "application/pdf":
            import fitz  # pymupdf
            doc = fitz.open(stream=data, filetype="pdf")
            return "\n".join(page.get_text() for page in doc)

        if "wordprocessingml" in content_type or content_type == "application/msword":
            from docx import Document
            doc = Document(io.BytesIO(data))
            return "\n".join(p.text for p in doc.paragraphs if p.text.strip())

        if "spreadsheetml" in content_type:
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
            rows: list[str] = []
            for sheet in wb.worksheets:
                for row in sheet.iter_rows(values_only=True):
                    cells = [str(c) for c in row if c is not None]
                    if cells:
                        rows.append(" | ".join(cells))
            return "\n".join(rows)

        return ""

    # ── Parsing + HTML stripping ─────────────────────────────────────────────

    @staticmethod
    def _parse(raw: dict) -> Email:
        from_info = raw.get("from", {}).get("emailAddress", {})
        sender = f"{from_info.get('name', '')} <{from_info.get('address', '')}>".strip()
        subject = raw.get("subject") or "(nessun oggetto)"
        date = (raw.get("receivedDateTime") or "")[:10]
        preview = raw.get("bodyPreview") or ""

        # Corpo completo: estrai da body.content e strippa l'HTML
        body_obj = raw.get("body", {})
        body_content = body_obj.get("content", "")
        body_type = body_obj.get("contentType", "text")

        if body_type == "html" and body_content:
            body_text = GraphClient._strip_html(body_content)
        else:
            body_text = body_content

        # Fallback: se body è vuoto usa il preview
        if not body_text.strip():
            body_text = preview

        return Email(
            id=raw["id"],
            sender=sender,
            subject=subject,
            date=date,
            body=body_text,
            preview=preview,
            conversation_id=raw.get("conversationId"),
        )

    @staticmethod
    def _strip_html(html: str) -> str:
        """
        Converte HTML in testo pulito preservando la struttura logica.
        Rimuove script, style, tag nascosti. Normalizza spazi bianchi.
        """
        soup = BeautifulSoup(html, "html.parser")

        # Rimuovi script, style, commenti
        for tag in soup(["script", "style", "head"]):
            tag.decompose()

        # Aggiungi newline prima di blocchi strutturali
        for br in soup.find_all("br"):
            br.replace_with("\n")
        for tag_name in ["p", "div", "tr", "li", "h1", "h2", "h3", "h4", "h5", "h6"]:
            for tag in soup.find_all(tag_name):
                tag.insert_before("\n")
                tag.insert_after("\n")

        text = soup.get_text()

        # Normalizza: collassa righe vuote multiple in una sola
        text = re.sub(r"\n\s*\n+", "\n\n", text)
        # Rimuovi spazi iniziali/finali per riga
        lines = [line.strip() for line in text.splitlines()]
        text = "\n".join(lines)
        # Rimuovi spazi iniziali/finali globali
        return text.strip()
